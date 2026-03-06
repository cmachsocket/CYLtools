from docx import Document
from openpyxl import load_workbook
import csv

rezult=["" ,"" ,""]

rezult[1] = "团建联络员打分结果_自动评估.csv"
rezult[2] = '团建联络员工作记录表_填写结果.csv'
documentPath = '团建联络员工作记录表.docx'
workbookPath = '团建联络员打分表.xlsx'
comments = []
scores = []

def read_csv(direction):
    with open(rezult[direction], mode='r', encoding='utf-8') as file:
    # 创建 csv.reader 对象
        csv_reader = csv.reader(file)
        # 跳过表头后再读取数据
        next(csv_reader, None)
        for row in csv_reader:
            for cell in row:
                if direction == 1:
                    scores.append(cell)
                else:
                    comments.append(cell)



document = Document( documentPath )
workbook = load_workbook( workbookPath)	

def copy_to_document():
    headflag = 1
    table = document.tables[0]
    exhausted = False
    for row in table.rows:
        if(headflag == 1):
            headflag = 0
            continue
        rowheadflag = 1
        for cell in row.cells:
            if(len(comments) == 0):
                exhausted = True
                break
            if(rowheadflag == 1):
                rowheadflag = 0
                continue
            cell.text = comments.pop(0)
        if exhausted:
            break
    document.save( documentPath )

def copy_to_workbook():
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Workbook has no active worksheet")

    score_iter = iter(scores)
    exhausted = False
    for i in sheet.iter_rows(min_row=2, max_row=8, min_col=1, max_col=5):
        for j in i:
            value = next(score_iter, None)
            if value is None:
                exhausted = True
                break
            j.value = value
        if exhausted:
            break
    workbook.save( workbookPath )

def main():
    read_csv(1)
    read_csv(2)
    print ("Comments:", comments)
    print ("Scores:", scores)
    copy_to_document()
    copy_to_workbook()

if __name__ == "__main__":
    main()