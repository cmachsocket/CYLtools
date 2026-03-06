#!/usr/bin/env python3
"""Convert an Excel (.xlsx) file to CSV.

Default behavior:
- Input:  团建联络员打分表.xlsx
- Output: 团建联络员打分表.csv
- Sheet:  first worksheet

Usage:
    python xlsx_to_csv.py
    python xlsx_to_csv.py --input 团建联络员打分表.xlsx
    python xlsx_to_csv.py --input 团建联络员打分表.xlsx --output output.csv
    python xlsx_to_csv.py --sheet "Sheet1"

Requirements:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert .xlsx to .csv")
    parser.add_argument(
        "--input",
        default="团建联络员打分表.xlsx",
        help="Input .xlsx file path (default: 团建联络员打分表.xlsx)",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .csv file path (default: same name as input, .csv)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name to export (default: first worksheet)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output file if it already exists.",
    )
    return parser.parse_args()


def build_output_path(input_path: Path, output_arg: str | None) -> Path:
    if output_arg:
        return Path(output_arg).expanduser().resolve()
    return input_path.with_suffix(".csv")


def load_sheet(input_path: Path, sheet_name: str | None):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install it with: pip install openpyxl"
        ) from exc

    wb = load_workbook(filename=input_path, data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            wb.close()
            raise RuntimeError(
                f"Sheet not found: {sheet_name}. Available sheets: {available}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    return wb, ws


def value_to_text(value) -> str:
    if value is None:
        return ""
    return str(value)


def write_csv(ws, output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([value_to_text(v) for v in row])


def main() -> int:
    args = parse_args()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists() or not input_path.is_file():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        return 1

    if input_path.suffix.lower() != ".xlsx":
        print("Error: input file must be an .xlsx file.", file=sys.stderr)
        return 1

    output_path = build_output_path(input_path, args.output)

    if output_path.exists() and not args.overwrite:
        print(
            f"Error: output already exists: {output_path}\\n"
            "Use --overwrite to replace it.",
            file=sys.stderr,
        )
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb, ws = load_sheet(input_path, args.sheet)
        write_csv(ws, output_path)
        wb.close()
    except RuntimeError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # noqa: BLE001
        print(f"Error: conversion failed ({exc})", file=sys.stderr)
        return 2

    print(f"OK: {input_path} [{ws.title}] -> {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
